"""
Excel Parser — Reads .xlsx files and builds a validated DataContext.

Handles:
- Multiple sheets with independent column mappings
- Type coercion (string, number, date, boolean)
- Required field validation (fail-fast before execution)
- Auto-detection of column names from header row
- Row filtering and range selection

Uses openpyxl for .xlsx parsing.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from data_provider.data_models import (
    DataContext,
    DataRow,
    DataSourceConfig,
    ExcelSheetConfig,
    ColumnMapping,
    IterationConfig,
    IterationMode,
    RowFilter,
)

logger = logging.getLogger("qa.data_provider.excel")


class ExcelParser:
    """
    Parses an Excel workbook and produces a validated DataContext.

    Workflow:
    1. Open workbook with openpyxl
    2. For each configured sheet:
       a. Read header row → build column map
       b. Read data rows → coerce types
       c. Validate required fields
       d. Apply filters if configured
    3. Return DataContext with all rows ready for template injection
    """

    def __init__(self, file_path: str, config: DataSourceConfig):
        self.file_path = Path(file_path)
        self.config = config
        self._workbook = None

    def parse(self) -> DataContext:
        """Parse the Excel file and return a validated DataContext."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Install openpyxl: pip install openpyxl")

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        logger.info(f"Parsing Excel: {self.file_path}")

        password = None
        if self.config.excel and self.config.excel.password:
            password = self.config.excel.password

        self._workbook = openpyxl.load_workbook(
            str(self.file_path),
            read_only=True,
            data_only=True,  # Resolve formulas to values
        )

        sheets_data: dict[str, list[DataRow]] = {}
        total_rows = 0
        validation_errors: list[str] = []

        excel_cfg = self.config.excel
        if not excel_cfg:
            # Auto-detect: read first sheet with default settings
            sheet_name = self._workbook.sheetnames[0]
            excel_cfg_sheets = [ExcelSheetConfig(sheet_name=sheet_name)]
        else:
            excel_cfg_sheets = excel_cfg.sheets

        for sheet_cfg in excel_cfg_sheets:
            try:
                rows, errors = self._parse_sheet(sheet_cfg)
                sheets_data[sheet_cfg.sheet_name] = rows
                total_rows += len(rows)
                validation_errors.extend(errors)

                logger.info(
                    f"  Sheet '{sheet_cfg.sheet_name}': {len(rows)} data rows, "
                    f"{len(errors)} validation errors"
                )
            except Exception as exc:
                error_msg = f"Error parsing sheet '{sheet_cfg.sheet_name}': {exc}"
                validation_errors.append(error_msg)
                logger.error(error_msg)

        self._workbook.close()

        # Apply iteration filters
        if self.config.iteration:
            sheets_data = self._apply_iteration(sheets_data, self.config.iteration)
            total_rows = sum(len(rows) for rows in sheets_data.values())

        context = DataContext(
            source_id=self.config.source_id,
            source_file=str(self.file_path),
            loaded_at=datetime.utcnow().isoformat(),
            sheets=sheets_data,
            total_rows=total_rows,
            validation_errors=validation_errors,
        )

        if validation_errors:
            logger.warning(
                f"Data loaded with {len(validation_errors)} validation errors. "
                f"Review before execution."
            )

        return context

    def _parse_sheet(
        self, sheet_cfg: ExcelSheetConfig
    ) -> tuple[list[DataRow], list[str]]:
        """Parse a single sheet and return (rows, validation_errors)."""
        ws = self._workbook[sheet_cfg.sheet_name]
        errors: list[str] = []

        # Build column mapping
        col_map = self._build_column_map(ws, sheet_cfg)

        # Read data rows
        rows: list[DataRow] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=sheet_cfg.data_start_row), start=1):
            # Check if row is completely empty (end of data)
            if all(cell.value is None for cell in row):
                continue

            values: dict[str, Any] = {}
            row_errors: list[str] = []
            excel_row_num = sheet_cfg.data_start_row + row_idx - 1

            for mapping in col_map:
                col_idx = self._col_letter_to_index(mapping.column)
                if col_idx >= len(row):
                    cell_value = None
                else:
                    cell_value = row[col_idx].value

                # Handle empty cells
                if cell_value is None or str(cell_value).strip() == "":
                    if mapping.required:
                        row_errors.append(
                            f"Row {excel_row_num}: Required field '{mapping.variable_name}' "
                            f"(column {mapping.column}) is empty"
                        )
                        cell_value = None
                    else:
                        cell_value = mapping.default_value
                else:
                    # Type coercion
                    cell_value = self._coerce_type(
                        cell_value, mapping, excel_row_num, row_errors
                    )

                values[mapping.variable_name] = cell_value

            errors.extend(row_errors)
            rows.append(DataRow(
                row_index=excel_row_num,
                values=values,
                sheet_name=sheet_cfg.sheet_name,
            ))

        return rows, errors

    def _build_column_map(
        self, ws, sheet_cfg: ExcelSheetConfig
    ) -> list[ColumnMapping]:
        """
        Build column mappings. If explicit mappings exist, use them.
        Otherwise, auto-detect from header row.
        """
        if sheet_cfg.column_mappings:
            return sheet_cfg.column_mappings

        # Auto-detect from header row
        header_row = list(ws.iter_rows(
            min_row=sheet_cfg.header_row,
            max_row=sheet_cfg.header_row,
        ))[0]

        mappings = []
        for idx, cell in enumerate(header_row):
            if cell.value is not None:
                col_letter = self._index_to_col_letter(idx)
                var_name = (
                    str(cell.value)
                    .strip()
                    .lower()
                    .replace(" ", "_")
                    .replace("-", "_")
                    .replace(".", "_")
                )
                mappings.append(ColumnMapping(
                    column=col_letter,
                    variable_name=var_name,
                    required=False,  # Auto-detected columns default to optional
                ))

        logger.debug(
            f"Auto-detected {len(mappings)} columns: "
            f"{[m.variable_name for m in mappings]}"
        )
        return mappings

    def _coerce_type(
        self,
        value: Any,
        mapping: ColumnMapping,
        row_num: int,
        errors: list[str],
    ) -> Any:
        """Coerce a cell value to the expected type."""
        try:
            if mapping.data_type == "string":
                return str(value).strip()

            elif mapping.data_type == "number":
                if isinstance(value, (int, float)):
                    return value
                cleaned = str(value).replace(",", "").strip()
                if "." in cleaned:
                    return float(cleaned)
                return int(cleaned)

            elif mapping.data_type == "date":
                if isinstance(value, datetime):
                    fmt = mapping.date_format or "%Y-%m-%d"
                    return value.strftime(fmt)
                if mapping.date_format:
                    dt = datetime.strptime(str(value), mapping.date_format)
                    return dt.strftime(mapping.date_format)
                return str(value)

            elif mapping.data_type == "boolean":
                if isinstance(value, bool):
                    return value
                return str(value).lower() in ("true", "yes", "1", "si", "sí")

            return str(value)

        except (ValueError, TypeError) as exc:
            errors.append(
                f"Row {row_num}: Cannot convert '{value}' to {mapping.data_type} "
                f"for field '{mapping.variable_name}' (column {mapping.column}): {exc}"
            )
            return str(value)

    def _apply_iteration(
        self,
        sheets_data: dict[str, list[DataRow]],
        iteration: IterationConfig,
    ) -> dict[str, list[DataRow]]:
        """Apply iteration config to filter/select rows."""
        sheet_name = iteration.sheet_name
        if sheet_name not in sheets_data:
            return sheets_data

        rows = sheets_data[sheet_name]

        if iteration.mode == IterationMode.SINGLE_ROW and iteration.specific_row:
            target_row = iteration.specific_row
            rows = [r for r in rows if r.row_index == target_row]

        elif iteration.mode == IterationMode.ROW_RANGE:
            start = iteration.row_start or 0
            end = iteration.row_end or float("inf")
            rows = [r for r in rows if start <= r.row_index <= end]

        elif iteration.mode == IterationMode.FILTERED:
            rows = self._apply_filters(rows, iteration.filters)

        # Safety limit
        if iteration.max_rows and len(rows) > iteration.max_rows:
            logger.warning(
                f"Row limit reached: {len(rows)} rows exceeds max_rows={iteration.max_rows}. "
                f"Truncating."
            )
            rows = rows[:iteration.max_rows]

        sheets_data[sheet_name] = rows
        return sheets_data

    def _apply_filters(
        self, rows: list[DataRow], filters: list[RowFilter]
    ) -> list[DataRow]:
        """Apply filter conditions to rows."""
        filtered = rows
        for f in filters:
            col_name = f.column
            filtered = [
                r for r in filtered
                if self._matches_filter(r.values.get(col_name), f)
            ]
        return filtered

    @staticmethod
    def _matches_filter(cell_value: Any, f: RowFilter) -> bool:
        """Check if a cell value matches a filter condition."""
        if f.operator == "not_empty":
            return cell_value is not None and str(cell_value).strip() != ""

        if cell_value is None:
            return False

        cell_str = str(cell_value)
        compare = f.value or ""

        if f.operator == "equals":
            return cell_str == compare
        elif f.operator == "contains":
            return compare in cell_str
        elif f.operator == "starts_with":
            return cell_str.startswith(compare)
        elif f.operator == "gt":
            try:
                return float(cell_str) > float(compare)
            except ValueError:
                return False
        elif f.operator == "lt":
            try:
                return float(cell_str) < float(compare)
            except ValueError:
                return False

        return True

    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        """Convert Excel column letter(s) to 0-based index. A=0, B=1, ..., AA=26."""
        result = 0
        for char in col.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1

    @staticmethod
    def _index_to_col_letter(idx: int) -> str:
        """Convert 0-based index to Excel column letter."""
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result
