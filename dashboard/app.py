"""Dashboard web app for JDE repetitive task automation.

Workflow:
1. Start Browser → login to JDE
2. Load Excel data → preview rows
3. Execute iterations → process each row
4. View report → see results
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from dashboard.session_manager import SessionManager
from data_provider.template_resolver import TemplateResolver
from models.schemas import TestSuiteRequest
from reports.html_report import generate_report
from models.schemas import (
    EngineType,
    SuiteResult,
    TestResult,
    TestStatus,
    StepResult,
)

# ---------------------------------------------------------------------------
# App state
# ---------------------------------------------------------------------------

_session = SessionManager()
_suite_request: Optional[TestSuiteRequest] = None
_execution_results: list[dict] = []
_row_paths: dict[int, str] = {}  # row_index → path name (full|a|b)
_report_groups: list[dict] = []  # per-column report groups ready for run_jde_full()
_login_completed: bool = False


# ---------------------------------------------------------------------------
# Path detection — choose which JSON to run for each row
# ---------------------------------------------------------------------------

PATH_TO_JSON: dict[str, str] = {
    "full": "tests/test_cases/jde_full.json",
    "a":    "tests/test_cases/jde_a_path.json",
    "b":    "tests/test_cases/jde_b_path.json",
}

# Excel layout — new format exported directly from JDE.
#
#   Row 1:   (blank or free-form header)
#   Row 2:   Object Name (A/B) + New Version Title per report column
#   Row 3:   "Copy from" label (A) + Current Version per report column
#   Row 4:   "DS Field" (A), "DATA SELECTION" (B) + New Version per report column
#   Row 5+:  Left Operand (A), Comparison (B) + New DS value per report column
#            ... until a separator row whose column A == "PO Tab" ...
#            then Processing Options: Tab (A), Option Number (B) + New value
#            per report column.
#
# Each column from C onward is one report iteration. Data Selections are
# the rows above the "PO Tab" separator; Processing Options are the rows
# below it. Column A is the field/tab, B is the comparison/option number.
JDE_META_ROW_TITLE = 2
JDE_META_ROW_CURRENT = 3
JDE_META_ROW_NEW = 4
JDE_DATA_START_ROW = 5
JDE_FIRST_REPORT_COL = 3  # column C
JDE_PO_SEPARATOR = "po tab"  # column-A marker (compared lower-cased)


def _extract_object_name(cell_values: list) -> str:
    """Pull the Object Name (App/Report) out of the first rows of the sheet.

    We accept any variation like:
        "Object Name: R4210IC"
        "Object Name : R4210IC"
        "R4210IC"                     (fallback — bare token)
    """
    for raw in cell_values:
        if raw is None:
            continue
        s = str(raw).strip()
        if not s:
            continue
        # Look for "Object Name: <TOKEN>"
        m = re.search(
            r"object\s*name\s*[:\-]?\s*([A-Za-z0-9_]+)",
            s,
            re.IGNORECASE,
        )
        if m:
            token = m.group(1).strip()
            if token and token.upper()[0] in ("R", "P"):
                return token
        # Otherwise: whole cell IS the token
        if re.match(r"^[RP][A-Za-z0-9_]+$", s):
            return s
    return ""


def _clean_left_operand(raw: str) -> str:
    """Normalize a JDE data-selection field name into a plain option label.

    The new Excel export names data selections like:

        "And BC Line Type (F411)(LNTY)"  ->  "Line Type"
        "BC Line Type (F411)(LNTY)"      ->  "Line Type"

    We strip, in order:
      - a leading boolean operator ("And"/"Or"), which is optional,
      - the leading two-letter section code that follows it ("BC"),
      - any parenthesized data-dictionary codes like "(F411)(LNTY)".

    The cleaned label is what we use to look up a Left Operand option.
    """
    if not raw:
        return ""
    s = str(raw)
    # Drop parenthesized codes anywhere, e.g. "(F411)(LNTY)"
    s = re.sub(r"\([^)]*\)", " ", s)
    # Drop a leading boolean operator ("And"/"Or"), if present
    s = re.sub(r"^\s*(?:and|or)\b\s*", "", s, flags=re.IGNORECASE)
    # Drop the leading two-letter section code (e.g. "BC")
    s = re.sub(r"^\s*[A-Za-z]{2}\b\s*", "", s)
    # Collapse leftover whitespace
    return re.sub(r"\s+", " ", s).strip()


def _po_label_first_segment(text: str) -> str:
    """Keep only the field-name portion of a Processing Option label (column B).

    JDE exports pack the option's help text after its name, separated by a run
    of 3+ spaces, e.g.:

        "5.  Prevent Next Status Update    Blank = Update next status  1 = ..."

    Only the leading segment ("5.  Prevent Next Status Update") names the field
    used to locate its text box, so we split on the first run of 3 or more
    spaces and keep the first part. Shorter gaps (like the "5.  " after the
    number) are preserved.
    """
    return re.split(r" {3,}", str(text or "").strip(), maxsplit=1)[0].strip()


# ---------------------------------------------------------------------------
# Data Selection value behavior
# ---------------------------------------------------------------------------
#
# Every extracted Data Selection value is classified into an edit *behavior*:
#
#   "remove"    → delete the matching row (REMOVE and "Blank" both map here)
#   "zero"      → select the "Zero" option in the Right Operand combo box
#   "null"      → select the "Null" option in the Right Operand combo box
#   "datetoday" → select the "DateToday [SL]" option in the Right Operand combo
#                 box (Excel value "SL DateToday")
#   "literal"   → a concrete value written via the Literal editor (default)
#
# The required data *type* for a literal is NOT inferred from the Left Operand
# name — it is determined at execution time from JDE's active Literal tab
# (Single Value / Range of Values / List of Values). See the executor's
# detect_active_literal_tab / _literal_tab_type_error.


def classify_ds_behavior(value: str) -> str:
    """Classify an extracted Data Selection value into an edit behavior.

    The team wraps the sentinel keywords in angle brackets in the new Excel
    template (``<Zero>``, ``<Null>``); a single surrounding ``<...>`` pair is
    stripped before matching, so both the new bracketed form and the legacy
    bare form (``Zero``/``Null``) classify identically.
    """
    v = str(value or "").strip()
    # Strip one surrounding pair of angle brackets: "<Zero>" -> "Zero".
    token = re.sub(r"^<\s*(.+?)\s*>$", r"\1", v).strip()
    low = token.lower()
    if token.upper() == "REMOVE" or low == "blank":
        return "remove"
    if low == "zero":
        return "zero"
    if low == "null":
        return "null"
    # "SL DateToday" → select the "DateToday [SL]" Right Operand option.
    if re.sub(r"\s+", " ", low) == "sl datetoday":
        return "datetoday"
    return "literal"


def parse_jde_excel_export(file_path: str, sheet_name: str) -> tuple[list[dict], list[dict]]:
    """Parse the JDE-exported Excel file into report groups.

    Returns (report_groups, skipped) where each report_group has the same
    shape run_jde_full expects: {report, data_selections, processing_options}.
    Processing Options come from the rows below the "PO Tab" separator row.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        # Read the first 4 rows in full (for Object Name detection + metadata)
        header_cells: list = []
        rows_by_index: dict[int, list] = {}
        for row_index, row in enumerate(
            ws.iter_rows(min_row=1, max_row=JDE_META_ROW_NEW, values_only=True), start=1
        ):
            rows_by_index[row_index] = list(row)
            for cell in row:
                header_cells.append(cell)

        app_report = _extract_object_name(header_cells)

        # Metadata rows
        row_title = rows_by_index.get(JDE_META_ROW_TITLE, [])
        row_current = rows_by_index.get(JDE_META_ROW_CURRENT, [])
        row_new = rows_by_index.get(JDE_META_ROW_NEW, [])

        # Rows from row 5 down are split into two sections by a separator row
        # whose column A holds the constant "PO Tab":
        #
        #   Data Selection rows (above the separator):
        #     A = Left Operand, B = Comparison, C+ = New value per report
        #   Processing Option rows (below the separator):
        #     A = Tab, B = option label, C+ = New value per report
        #
        # Column B is the option's full label text (e.g. "1. Sales Order
        # Entry (P4210)") — the executor searches it in JDE to locate the
        # text box to fill. (In the previous format Tab/Option Number/New
        # Value came from columns I/J/K; they now live in A/B/C+.)
        ds_rows: list[dict] = []
        po_rows: list[dict] = []
        in_po_section = False
        for row_index, row in enumerate(
            ws.iter_rows(min_row=JDE_DATA_START_ROW, values_only=True),
            start=JDE_DATA_START_ROW,
        ):
            row = list(row)
            col_a = row[0] if len(row) > 0 else None
            a_str = str(col_a).strip() if col_a is not None else ""

            # Separator row → everything below belongs to Processing Options.
            if a_str.lower() == JDE_PO_SEPARATOR:
                in_po_section = True
                continue
            if not a_str:
                continue

            col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if in_po_section:
                po_rows.append({
                    "row_index": row_index,
                    "row": row,
                    "tab": a_str,
                    # Keep only the field name; drop trailing help text that
                    # JDE separates with 3+ spaces.
                    "option_label": _po_label_first_segment(col_b),
                })
            else:
                left_operand = _clean_left_operand(a_str)
                ds_rows.append({
                    "row_index": row_index,
                    "row": row,
                    "left_operand": left_operand,
                    "comparison": col_b,
                })

        # Determine how many report columns we have (columns C..N).
        # A column is considered "present" if any of Row 2/3/4 has data.
        # Grow the search up to the widest row.
        max_len = max(
            len(row_title), len(row_current), len(row_new),
            *[len(r["row"]) for r in ds_rows],
            *[len(r["row"]) for r in po_rows],
            JDE_FIRST_REPORT_COL,
        )
        report_groups: list[dict] = []
        skipped: list[dict] = []

        # Excel columns are 1-indexed; column C = index 2 in a zero-based list
        for col_idx0 in range(JDE_FIRST_REPORT_COL - 1, max_len):
            col_letter = _col_letter(col_idx0 + 1)
            title = _cell(row_title, col_idx0)
            current = _cell(row_current, col_idx0)
            new_ver = _cell(row_new, col_idx0)

            # Row 4 (New Version) decides the mode PER report column, regardless
            # of the Column A / Row 3 "Copy from" label (which is ignored):
            #   text  → copy current_version into this new version.
            #   empty → edit current_version in place.
            is_copy_mode = bool(new_ver)

            # Skip completely empty columns
            if not any([title, current, new_ver]):
                # Also check if this column has ANY DS or PO values — if it
                # does, something is off. Otherwise just skip.
                has_values = any(_cell(r["row"], col_idx0) for r in ds_rows) or \
                    any(_cell(r["row"], col_idx0) for r in po_rows)
                if not has_values:
                    continue

            # Both modes need the version to open (Row 3). Copy mode also needs
            # the target name (Row 4), but that is what selected copy mode, so
            # its presence is already guaranteed here.
            if not current:
                skipped.append({
                    "row": col_letter,
                    "app_report": app_report,
                    "reason": (
                        f"Column {col_letter} missing "
                        f"{'current version' if is_copy_mode else 'version to edit'} (Row 3)"
                    ),
                })
                continue

            if app_report and not app_report.upper().startswith(("R", "P")):
                skipped.append({
                    "row": col_letter,
                    "app_report": app_report,
                    "reason": "App/Report must start with 'R' or 'P'",
                })
                continue

            # Collect data selections for this report column
            data_selections: list[dict] = []
            # Occurrence is counted PER COLUMN among the entries actually filled
            # here (not globally across the sheet): each report column is a
            # distinct JDE version, so its Nth filled "Order Company" maps to
            # the Nth "Order Company" row in that version's grid. Counting
            # globally would let a row left blank in THIS column still consume
            # an ordinal, so a lone value would look for a non-existent Nth row.
            col_occurrence: dict[str, int] = {}
            for r in ds_rows:
                val = _cell(r["row"], col_idx0)
                if val is None or not str(val).strip():
                    continue
                data_new = str(val).strip()
                lo_key = r["left_operand"].lower()
                col_occurrence[lo_key] = col_occurrence.get(lo_key, 0) + 1

                # Classify the edit behavior (remove / zero / null / literal).
                # The literal data type is verified at execution time from
                # JDE's active tab, not from the Left Operand name.
                data_selections.append({
                    "left_operand": r["left_operand"],
                    "comparison": r["comparison"],
                    "data_new": data_new,
                    "behavior": classify_ds_behavior(data_new),
                    "occurrence": col_occurrence[lo_key],
                    "_source_row": r["row_index"],
                })

            # Collect processing options for this report column. Tab = col A,
            # New Value = this report's column. Values are placed POSITIONALLY:
            # the Nth row of a tab fills the Nth text box in that JDE tab, so a
            # blank cell simply skips (leaves untouched) that text box. `position`
            # (0-based, counting blank rows) is the text-box index within the tab.
            processing_options: list[dict] = []
            tab_pos: dict[str, int] = {}
            for r in po_rows:
                tab = r["tab"]
                pos = tab_pos.get(tab, 0)
                tab_pos[tab] = pos + 1  # every row consumes a text-box slot
                val = _cell(r["row"], col_idx0)
                if val is None or not str(val).strip():
                    continue  # blank → skip this text box (position already used)
                processing_options.append({
                    "tab": tab,
                    "option_label": r["option_label"],
                    "processing_new": str(val).strip(),
                    "position": pos,
                    "_source_row": r["row_index"],
                })

            report_groups.append({
                "row_index": col_letter,  # keep letter for the UI preview
                "report": {
                    "app_report": app_report,
                    "current_version": str(current).strip(),
                    "new_version": str(new_ver).strip() if new_ver else "",
                    "new_version_title": str(title).strip() if title else "",
                    # True → copy current_version into new_version (default
                    # workflow). False → edit current_version in place.
                    "copy_version": is_copy_mode,
                },
                "data_selections": data_selections,
                "processing_options": processing_options,
            })

        return report_groups, skipped
    finally:
        wb.close()


_DATE_MDY = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$")
_DATE_ISO = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})(?:[ T]00:00:00)?$")


def _normalize_date_string(s: str) -> str:
    """Reformat a date-looking string to JDE's DD/MM/YY.

    The Excel export uses US order MM/DD/YYYY (07/26/2026); JDE expects
    DD/MM/YY (26/07/26). Also accepts ISO (2026-07-26, optionally with a
    midnight time). Returns *s* unchanged if it isn't a valid date.
    """
    m = _DATE_MDY.match(s)
    if m:
        mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    else:
        m = _DATE_ISO.match(s)
        if not m:
            return s
        yyyy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime(yyyy, mm, dd).strftime("%d/%m/%y")
    except ValueError:
        return s


def _cell(row: list, idx0: int):
    """Safely read row[idx0], return None if out of range or empty-ish.

    Date cells are normalized to JDE's DD/MM/YY format. openpyxl returns real
    Excel dates as datetime objects whose str() is an ISO
    "2026-07-26 00:00:00" — JDE mis-parses that into e.g. "7-26-2026". The Excel
    export writes dates as MM/DD/YYYY, so we format date objects and reshape
    date-looking text to DD/MM/YY so JDE stores the date correctly.
    """
    if idx0 >= len(row):
        return None
    v = row[idx0]
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%y")
    s = str(v).strip()
    if not s:
        return None
    return _normalize_date_string(s)


def _col_letter(col_num: int) -> str:
    """1 → 'A', 27 → 'AA', ..."""
    letters = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def _build_login_suite() -> "TestSuiteRequest":
    """Build the login TestSuiteRequest from login_assert.json, overriding URL
    and credentials from the environment. Reused by the /login endpoint and by
    the automatic re-login inside execute_all when the JDE session drops."""
    login_path = "tests/test_cases/login_assert.json"
    if not Path(login_path).exists():
        raise FileNotFoundError(f"Login suite not found: {login_path}")

    login_raw = json.loads(Path(login_path).read_text(encoding="utf-8"))
    login_raw.pop("_data_source", None)  # login suite has no data source

    jde_url = os.getenv("JDE_URL", "").strip()
    jde_user = os.getenv("JDE_USERNAME", "").strip()
    jde_pass = os.getenv("JDE_PASSWORD", "").strip()

    for tc in login_raw.get("test_cases", []):
        if jde_url:
            tc["base_url"] = jde_url
        for step in tc.get("steps", []):
            action = step.get("action")
            name = (step.get("name") or "").lower()
            if action == "navigate" and jde_url:
                step.setdefault("data", {})["value"] = jde_url
            elif action == "type" and "user" in name and jde_user:
                step.setdefault("data", {})["value"] = jde_user
            elif action == "type" and "password" in name and jde_pass:
                step.setdefault("data", {})["value"] = jde_pass

    login_suite = TestSuiteRequest(**login_raw)
    login_suite.headless = False
    return login_suite


def _to_modify_group(group: dict) -> dict:
    """Return a copy of *group* forced into edit/modify mode on the version that
    a dropped-session iteration was creating.

    When JDE logs out mid-iteration a copy may already have been created, so
    re-running in copy mode would fail ("version already exists"). Instead we
    open the New Version (falling back to Current Version) and rewrite its Data
    Selection / Processing Options in place.
    """
    report = dict(group.get("report", {}))
    target = report.get("new_version") or report.get("current_version")
    report["current_version"] = target
    report["new_version"] = ""          # empty New Version → edit mode
    report["copy_version"] = False
    new_group = dict(group)
    new_group["report"] = report
    return new_group


def create_dashboard_app() -> FastAPI:
    from contextlib import asynccontextmanager

    @asynccontextmanager
    async def lifespan(app: FastAPI):
        yield
        await _session.stop()

    app = FastAPI(
        title="JDE Automation Dashboard",
        version="1.0.0",
        lifespan=lifespan,
    )

    # --- Serve the frontend -------------------------------------------------

    @app.get("/", response_class=HTMLResponse)
    async def index():
        html_path = Path(__file__).parent / "index.html"
        return HTMLResponse(html_path.read_text(encoding="utf-8"))

    # --- Session endpoints --------------------------------------------------

    @app.post("/api/session/start")
    async def start_browser():
        """Launch the browser. Returns a clear error if launch fails."""
        try:
            return await _session.start_browser()
        except Exception as exc:
            import traceback
            tb = traceback.format_exc()
            _session.logger.error("Browser start failed: %s\n%s", exc, tb)
            # Make sure any partially-initialized state is cleaned up
            try:
                await _session.stop()
            except Exception:
                pass
            raise HTTPException(
                status_code=500,
                detail=f"Failed to launch browser: {type(exc).__name__}: {exc}",
            )

    @app.post("/api/session/login")
    async def login(request: Request):
        """Run login_assert.json — login only, no Excel data involved."""
        global _suite_request, _login_completed

        try:
            login_suite = _build_login_suite()
        except FileNotFoundError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

        _suite_request = login_suite
        result = await _session.run_login(login_suite)
        _login_completed = bool(result.get("logged_in"))
        return result

    @app.post("/api/session/stop")
    async def stop_browser():
        """Close the browser and reset all dashboard state."""
        global _suite_request
        global _execution_results, _row_paths, _report_groups, _login_completed
        result = await _session.stop()
        # Wipe server-side state so the dashboard is ready for a fresh run
        _suite_request = None
        _execution_results = []
        _row_paths = {}
        _report_groups = []
        _login_completed = False
        return result

    @app.get("/api/session/status")
    async def session_status():
        """Get current session state."""
        return {
            "browser_active": _session.is_active,
            "logged_in": _session.is_logged_in,
            "suite_loaded": _suite_request is not None,
            "data_loaded": bool(_report_groups),
            "data_rows": len(_report_groups),
            "executions_completed": len(_execution_results),
        }

    # --- Data endpoints -----------------------------------------------------

    @app.post("/api/data/sheets")
    async def list_sheets(file: UploadFile = File(...)):
        """Return the list of sheet names in the uploaded xlsx file.

        Used to populate the sheet-name combo box in the dashboard before
        the user picks which sheet to parse.
        """
        if not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files accepted")

        try:
            content = await file.read()
            if len(content) > 50 * 1024 * 1024:
                raise HTTPException(status_code=413, detail="File too large (max 50 MB)")
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to read upload: {exc}")

        # Open the workbook from the in-memory bytes and list sheet names
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not read xlsx: {exc}")

        return {"filename": file.filename, "sheets": sheets}

    @app.post("/api/data/upload")
    async def upload_excel(
        file: UploadFile = File(...),
        sheet_name: str = Form("Sheet1"),
    ):
        """Upload an xlsx file, save to temp, and parse it."""
        if not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files accepted")

        if not _login_completed:
            raise HTTPException(status_code=400, detail="Run Start Browser & Login first.")

        # Write the upload to a TEMP file (deleted after parsing) so the
        # run-folder doesn't fill up with xlsx copies the user doesn't want.
        import tempfile
        try:
            content = await file.read()
            if len(content) > 50 * 1024 * 1024:
                raise HTTPException(status_code=413, detail="File too large (max 50 MB)")
            tf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            try:
                tf.write(content)
                tf.flush()
                saved_path = Path(tf.name)
            finally:
                tf.close()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to save upload: {exc}")

        # Parse the JDE-exported workbook using the format-specific parser
        # (columns C..N are one report iteration each; metadata is in rows 2-4).
        try:
            groups, skipped_rows = parse_jde_excel_export(
                str(saved_path), sheet_name.strip() or "Sheet1"
            )
        except Exception as exc:
            import traceback
            err = f"{type(exc).__name__}: {exc}"
            _session.logger.error("Excel parse error: %s\n%s", err, traceback.format_exc())
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {err}")
        finally:
            # Always remove the temp xlsx — the data is already in memory
            try:
                saved_path.unlink(missing_ok=True)
            except Exception:
                pass

        global _report_groups
        _report_groups = groups

        # Log a per-column summary for debugging
        for g in _report_groups:
            _session.logger.info(
                "Report col %s: app=%s current=%s new=%s mode=%s DS=%d PO=%d",
                g["row_index"],
                g["report"].get("app_report"),
                g["report"].get("current_version"),
                g["report"].get("new_version"),
                "copy" if g["report"].get("copy_version", True) else "edit",
                len(g["data_selections"]),
                len(g["processing_options"]),
            )

        # Build a flat preview — one row per report group (column)
        preview = []
        for group in _report_groups:
            report = group["report"]
            preview.append({
                "_row": group["row_index"],
                "app_report": report.get("app_report", ""),
                "current_version": report.get("current_version", ""),
                "new_version": report.get("new_version", ""),
                "new_version_title": report.get("new_version_title", ""),
                "copy_version": report.get("copy_version", True),
                "data_selections_count": len(group["data_selections"]),
                "processing_options_count": len(group["processing_options"]),
            })

        return {
            "status": "success",
            "filename": file.filename,
            "rows": len(_report_groups),
            "skipped_rows": skipped_rows,
            "skipped_count": len(skipped_rows),
            "preview": preview,
        }

    @app.get("/api/data/preview")
    async def data_preview():
        """Get the loaded data preview."""
        if not _report_groups:
            raise HTTPException(status_code=400, detail="No data loaded")
        preview = []
        for g in _report_groups:
            r = g["report"]
            preview.append({
                "_row": g["row_index"],
                "app_report": r.get("app_report", ""),
                "current_version": r.get("current_version", ""),
                "new_version": r.get("new_version", ""),
                "new_version_title": r.get("new_version_title", ""),
                "copy_version": r.get("copy_version", True),
                "data_selections_count": len(g["data_selections"]),
                "processing_options_count": len(g["processing_options"]),
            })
        return {"rows": len(_report_groups), "preview": preview}

    # --- Execution endpoints ------------------------------------------------

    @app.post("/api/execute")
    async def execute_all():
        """Run iterations: each report group runs the JDE Full Path Python flow."""
        global _execution_results
        import time
        import asyncio as _asyncio
        from tests.test_jde_full import run_jde_full, is_session_active

        if not _session.is_logged_in:
            raise HTTPException(status_code=400, detail="Not logged in. Run login first.")
        if not _report_groups:
            raise HTTPException(status_code=400, detail="No data loaded. Load Excel first.")

        _execution_results = []
        total = len(_report_groups)
        page = _session._page  # the persistent logged-in page

        if page is None:
            raise HTTPException(status_code=500, detail="Browser page is not available")

        def _record(iter_num: int, grp: dict, status: str, error, duration_ms: float,
                    steps_raw: list, modified: bool):
            report = grp["report"]
            label = (f"{report.get('app_report', '?')} → "
                     f"{report.get('new_version') or report.get('current_version', '?')}")
            steps_dicts = []
            total_tokens = 0
            for s in steps_raw:
                total_tokens += getattr(s, "tokens_used", 0) or 0
                steps_dicts.append({
                    "step_id": s.step_id,
                    "name": s.name,
                    "action": s.action.value if hasattr(s.action, "value") else str(s.action),
                    "status": s.status.value if hasattr(s.status, "value") else str(s.status),
                    "duration_ms": round(s.duration_ms or 0),
                    "tokens_used": s.tokens_used or 0,
                    "error": s.error_message,
                    "selector": s.resolved_selector,
                    "started_at": s.started_at.isoformat() if s.started_at else None,
                    "finished_at": s.finished_at.isoformat() if s.finished_at else None,
                })
            # Surface a failure that produced no steps as a synthetic step.
            if not steps_dicts and status == "fail":
                steps_dicts.append({
                    "step_id": "S000", "name": "Iteration failed", "action": "custom",
                    "status": "fail", "duration_ms": round(duration_ms), "tokens_used": 0,
                    "error": error, "selector": None,
                })
            _execution_results.append({
                "iteration": iter_num,
                "total": total,
                "test_id": report.get("app_report", "?"),
                "name": label + (" [modified after re-login]" if modified else ""),
                "status": status,
                "duration_ms": round(duration_ms),
                "tokens": total_tokens,
                "screenshot": "",
                "modified": modified,
                "data_selections_count": len(grp["data_selections"]),
                "processing_options_count": len(grp["processing_options"]),
                "steps": steps_dicts,
            })

        # A very long batch (e.g. 50 versions) can outlive the JDE session. After
        # each iteration we check whether the "Sign Out" control is still visible;
        # if the session dropped, we re-login and re-run that same iteration in
        # MODIFY mode (rewriting its New Version in place, since a copy may have
        # been created before the logout), then continue with the rest.
        MAX_RELOGIN = int(os.getenv("JDE_MAX_RELOGIN", "5"))
        relogin_total = 0
        modify_indices: set[int] = set()  # iterations to retry in modify mode

        i = 0
        while i < total:
            group = _report_groups[i]
            iter_num = i + 1
            use_modify = i in modify_indices
            run_group = _to_modify_group(group) if use_modify else group
            report = run_group["report"]
            label = (f"{report.get('app_report', '?')} → "
                     f"{report.get('new_version') or report.get('current_version', '?')}")
            print(f"\n=== Iteration {iter_num}/{total}: {label}"
                  f"{' [modify after re-login]' if use_modify else ''} ===")
            print(f"   Data selections: {len(group['data_selections'])}, "
                  f"Processing options: {len(group['processing_options'])}")

            start = time.monotonic()
            steps_raw = []  # list of StepResult objects from the runner
            try:
                result = await run_jde_full(page, run_group)
                status = result.get("status", "fail")
                error = result.get("error")
                steps_raw = result.get("steps") or []
            except Exception as exc:
                import traceback
                _session.logger.error("Iteration %d crashed: %s\n%s", iter_num, exc, traceback.format_exc())
                status = "fail"
                error = f"Unhandled exception: {exc}"

            duration_ms = (time.monotonic() - start) * 1000

            # Did the JDE session drop during this iteration?
            session_ok = True
            try:
                session_ok = await is_session_active(page)
                if not session_ok:
                    await _asyncio.sleep(1.0)  # avoid a transient false negative
                    session_ok = await is_session_active(page)
            except Exception:
                session_ok = True  # never let the health check itself block a run

            if not session_ok and i not in modify_indices and relogin_total < MAX_RELOGIN:
                relogin_total += 1
                print(f"[exec] ⚠ Session lost during iteration {iter_num} — "
                      f"re-logging in (#{relogin_total}) and retrying in modify mode")
                _session.logger.warning(
                    "Session lost during iteration %d — re-login attempt #%d",
                    iter_num, relogin_total,
                )
                try:
                    relogin = await _session.run_login(_build_login_suite())
                except Exception as exc:
                    relogin = {"logged_in": False, "error": str(exc)}
                if relogin.get("logged_in"):
                    # Re-run THIS iteration in modify mode; don't record the
                    # failed attempt, don't advance.
                    modify_indices.add(i)
                    page = _session._page or page  # refresh in case it changed
                    continue
                # Re-login failed — record this one and stop; the rest can't run.
                _record(iter_num, run_group, "fail",
                        "JDE session logged out during this iteration and automatic "
                        f"re-login failed: {relogin.get('error') or 'login assert not confirmed'}",
                        duration_ms, steps_raw, use_modify)
                for j in range(i + 1, total):
                    _record(j + 1, _report_groups[j], "fail",
                            "Skipped: JDE session logged out and re-login failed",
                            0.0, [], False)
                break

            _record(iter_num, run_group, status, error, duration_ms, steps_raw, use_modify)
            i += 1

        # Generate report
        report_path = _generate_execution_report()

        passed = sum(1 for r in _execution_results if r["status"] == "pass")
        failed = total - passed

        return {
            "status": "completed",
            "total": total,
            "passed": passed,
            "failed": failed,
            "report_path": report_path,
            "results": _execution_results,
        }

    @app.get("/api/execute/results")
    async def get_results():
        """Get execution results."""
        if not _execution_results:
            return {"status": "no_results", "results": []}

        passed = sum(1 for r in _execution_results if r["status"] == "pass")
        return {
            "status": "completed",
            "total": len(_execution_results),
            "passed": passed,
            "failed": len(_execution_results) - passed,
            "results": _execution_results,
        }

    @app.get("/api/report")
    async def get_report():
        """Get the HTML report."""
        if not _execution_results:
            raise HTTPException(status_code=400, detail="No execution results")

        report_path = _generate_execution_report()
        html = Path(report_path).read_text(encoding="utf-8")
        return HTMLResponse(html)

    return app


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _generate_execution_report() -> str:
    """Build a SuiteResult from execution results and generate HTML report."""
    if not _suite_request or not _execution_results:
        return ""

    from models.schemas import ActionType
    from datetime import datetime as _dt

    def _parse_action(value: str) -> ActionType:
        try:
            return ActionType(value)
        except Exception:
            return ActionType.CUSTOM

    def _parse_ts(value):
        if not value:
            return None
        try:
            return _dt.fromisoformat(value)
        except Exception:
            return None

    test_results = []
    for r in _execution_results:
        steps = [
            StepResult(
                step_id=s["step_id"],
                name=s["name"],
                action=_parse_action(s.get("action", "custom")),
                status=s["status"],
                duration_ms=s.get("duration_ms", 0),
                tokens_used=s.get("tokens_used", 0),
                error_message=s.get("error"),
                resolved_selector=s.get("selector"),
                started_at=_parse_ts(s.get("started_at")),
                finished_at=_parse_ts(s.get("finished_at")),
            )
            for s in r.get("steps", [])
        ]
        test_results.append(TestResult(
            test_id=f"{r['test_id']}_iter{r['iteration']}",
            name=r["name"],
            status=r["status"],
            platform=_suite_request.test_cases[0].platform if _suite_request.test_cases else "generic_web",
            steps=steps,
            duration_ms=r.get("duration_ms", 0),
            total_tokens=r.get("tokens", 0),
        ))

    suite_result = SuiteResult(
        suite_id=_suite_request.suite_id,
        suite_name=_suite_request.suite_name,
        environment=_suite_request.environment,
        browser=_suite_request.browser,
        engine_type=EngineType.HYBRID,
        llm_provider=_suite_request.llm_provider,
        llm_model=_suite_request.llm_model,
        test_results=test_results,
        started_at=datetime.now(),
        finished_at=datetime.now(),
        total_duration_ms=sum(r.get("duration_ms", 0) for r in _execution_results),
        total_tokens=sum(r.get("tokens", 0) for r in _execution_results),
    )

    run_dir = _session.run_dir or Path("logs")
    return generate_report(suite_result, output_dir=str(run_dir), filename="report.html")


dashboard_app = create_dashboard_app()
